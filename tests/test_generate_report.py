import subprocess
import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.generate_report import (
    DEFAULT_EXCEL_OUTPUT,
    DEFAULT_HTML_OUTPUT,
    create_excel_report,
)
from scripts.web_report import TEMPLATE_PATH, create_html_report


class GenerateReportTests(unittest.TestCase):
    def test_browser_report_source_uses_report_layout_directory(self):
        self.assertEqual(TEMPLATE_PATH.parent.name, "report_layout")
        self.assertTrue(TEMPLATE_PATH.is_file())

    def test_generates_report_from_synthetic_excel_sources(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            temporary_path = Path(temporary_directory)
            company_path = temporary_path / "company_information.xlsx"
            sales_path = temporary_path / "regional_sales.xlsx"
            excel_output_path = (
                temporary_path
                / "regional_sales_reports"
                / "regional_sales_report.xlsx"
            )
            html_output_path = temporary_path / "outputs" / "site" / "index.html"
            browser_excel_path = (
                html_output_path.parent / "regional_sales_report.xlsx"
            )

            company_workbook = Workbook()
            company_sheet = company_workbook.active
            company_sheet.append(
                [
                    "company_id",
                    "company_name",
                    "GICS_sector",
                    "GICS_sub_industry",
                    "headquarters_loc",
                ]
            )
            company_sheet.append(
                [
                    "CMP0001",
                    "Aveden",
                    "Health Care",
                    "Biotechnology",
                    "Stockholm, Sweden",
                ]
            )
            company_workbook.save(company_path)

            sales_workbook = Workbook()
            sales_sheet = sales_workbook.active
            sales_sheet.append(
                ["company_id", "gareac", "salecs", "reporting_date"]
            )
            sales_sheet.append(
                ["CMP0001", "Europe", 2704.268, date(2023, 3, 31)]
            )
            sales_workbook.save(sales_path)

            result = subprocess.run(
                [
                    sys.executable,
                    "scripts/generate_report.py",
                    "2023-01-01",
                    "2023-03-31",
                    "--company-data",
                    str(company_path),
                    "--sales-data",
                    str(sales_path),
                    "--output",
                    str(excel_output_path),
                    "--html-output",
                    str(html_output_path),
                ],
                capture_output=True,
                text=True,
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertTrue(excel_output_path.exists())
            self.assertTrue(html_output_path.exists())
            self.assertTrue(browser_excel_path.exists())

            report_workbook = load_workbook(excel_output_path)
            report_sheet = report_workbook["Regional Sales Report"]
            self.assertEqual(report_sheet["A8"].value, "Aveden")
            self.assertEqual(report_sheet["D8"].value, "Europe")
            self.assertEqual(report_sheet["E8"].value, 2704.268)

            html = html_output_path.read_text(encoding="utf-8")
            self.assertIn("Aveden", html)
            self.assertIn("Europe", html)

    def test_default_outputs_separate_saved_and_published_reports(self):
        project_root = Path(__file__).resolve().parents[1]
        report_directory = project_root / "regional_sales_reports"
        site_directory = project_root / "outputs" / "site"

        self.assertEqual(
            DEFAULT_EXCEL_OUTPUT,
            report_directory / "regional_sales_report.xlsx",
        )
        self.assertEqual(
            DEFAULT_HTML_OUTPUT,
            site_directory / "index.html",
        )

    def test_rejects_end_date_before_start_date(self):
        result = subprocess.run(
            [
                sys.executable,
                "scripts/generate_report.py",
                "2023-12-31",
                "2023-10-01",
            ],
            capture_output=True,
            text=True,
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn(
            "start date must not be after end date",
            result.stderr.lower(),
        )

    def test_creates_excel_report(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            output_path = (
                Path(temporary_directory) / "regional_sales_report.xlsx"
            )
            html_output_path = Path(temporary_directory) / "index.html"

            result = subprocess.run(
                [
                    sys.executable,
                    "scripts/generate_report.py",
                    "2023-10-01",
                    "2023-12-31",
                    "--output",
                    str(output_path),
                    "--html-output",
                    str(html_output_path),
                ],
                capture_output=True,
                text=True,
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertTrue(output_path.exists())
            self.assertTrue(html_output_path.exists())

    def test_workbook_uses_one_report_sheet_and_one_filter(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            output_path = (
                Path(temporary_directory) / "regional_sales_report.xlsx"
            )

            create_excel_report(
                output_path,
                date(2023, 10, 1),
                date(2023, 12, 31),
                [
                    "company_name",
                    "GICS_sector",
                    "GICS_sub_industry",
                    "geographic_area_code",
                    "total_sales",
                ],
                [
                    (
                        "Example Company",
                        "Industrials",
                        "Business Services",
                        "USA",
                        1250.5,
                    )
                ],
            )

            workbook = load_workbook(output_path)

            self.assertEqual(
                workbook.sheetnames,
                ["Regional Sales Report"],
            )

            report_sheet = workbook["Regional Sales Report"]
            self.assertEqual(report_sheet["A1"].value, "Regional Sales Report")
            self.assertEqual(report_sheet.max_row, 8)
            self.assertEqual(len(report_sheet.tables), 1)
            self.assertIsNone(report_sheet.auto_filter.ref)

    def test_html_report_groups_companies_and_calculates_subtotals(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            output_path = Path(temporary_directory) / "index.html"

            create_html_report(
                output_path,
                date(2023, 10, 1),
                date(2023, 12, 31),
                [
                    "company_name",
                    "GICS_sector",
                    "GICS_sub_industry",
                    "geographic_area_code",
                    "total_sales",
                ],
                [
                    (
                        "Omnicom Group",
                        "Communication Services",
                        "Advertising",
                        "EUROPE",
                        10,
                    ),
                    (
                        "Omnicom Group",
                        "Communication Services",
                        "Advertising",
                        "USA",
                        15,
                    ),
                    (
                        "Other Company",
                        "Communication Services",
                        "Advertising",
                        "USA",
                        5,
                    ),
                ],
                "regional_sales_report.xlsx",
            )

            html = output_path.read_text(encoding="utf-8")

            self.assertEqual(html.count(">Omnicom Group<"), 1)
            self.assertIn("25.000", html)
            self.assertIn("30.000", html)
            self.assertIn('id="report-search"', html)
            self.assertIn('id="sector-filter"', html)
            self.assertIn("<details", html)
            self.assertIn('href="regional_sales_report.xlsx"', html)

    def test_html_report_escapes_values_and_handles_no_results(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            escaped_output = Path(temporary_directory) / "escaped.html"
            empty_output = Path(temporary_directory) / "empty.html"

            create_html_report(
                escaped_output,
                date(2023, 10, 1),
                date(2023, 12, 31),
                [],
                [
                    (
                        "Company <script>alert(1)</script>",
                        "Industrials",
                        "Services",
                        "USA",
                        1,
                    )
                ],
                "regional_sales_report.xlsx",
            )
            create_html_report(
                empty_output,
                date(2024, 1, 1),
                date(2024, 3, 31),
                [],
                [],
                "regional_sales_report.xlsx",
            )

            escaped_html = escaped_output.read_text(encoding="utf-8")
            empty_html = empty_output.read_text(encoding="utf-8")

            self.assertNotIn("<script>alert(1)</script>", escaped_html)
            self.assertIn(
                "Company &lt;script&gt;alert(1)&lt;/script&gt;",
                escaped_html,
            )
            self.assertIn("No report rows were found", empty_html)

    def test_workflow_saves_excel_report_and_deploys_pages(self):
        workflow = (
            Path(__file__).resolve().parents[1]
            / ".github"
            / "workflows"
            / "generate_report.yml"
        ).read_text(encoding="utf-8")

        self.assertIn("actions/configure-pages@v6", workflow)
        self.assertIn("actions/upload-pages-artifact@v5", workflow)
        self.assertNotIn("actions/upload-artifact@", workflow)
        self.assertIn("actions/deploy-pages@v5", workflow)
        self.assertIn(
            "--output regional_sales_reports/regional_sales_report.xlsx",
            workflow,
        )
        self.assertIn(
            "git add regional_sales_reports/regional_sales_report.xlsx",
            workflow,
        )
        self.assertIn('git commit -m "Update regional sales report"', workflow)
        self.assertIn("git push", workflow)

    def test_report_uses_sqlite_and_pandas_flow(self):
        project_root = Path(__file__).resolve().parents[1]
        script = (
            project_root / "scripts" / "generate_report.py"
        ).read_text(encoding="utf-8")
        requirements = (
            project_root / "requirements.txt"
        ).read_text(encoding="utf-8")

        self.assertIn("import sqlite3", script)
        self.assertIn("import pandas as pd", script)
        self.assertIn("sqlite3.connect", script)
        self.assertIn("regional_sales.sqlite", script)
        self.assertIn("pd.read_sql_query", script)
        self.assertIn("df.to_excel", script)
        self.assertNotIn("with connection", script)
        self.assertNotIn("import duckdb", script)
        self.assertIn("pandas", requirements)
        self.assertNotIn("duckdb", requirements.lower())


if __name__ == "__main__":
    unittest.main()
