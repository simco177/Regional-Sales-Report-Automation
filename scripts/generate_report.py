import argparse
import sqlite3
import shutil
from datetime import date, datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

if __package__:
    from scripts.web_report import create_html_report
else:
    from web_report import create_html_report


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATABASE_PATH = PROJECT_ROOT / "database" / "regional_sales.sqlite"
SITE_OUTPUT_DIRECTORY = PROJECT_ROOT / "outputs" / "site"
REPORT_OUTPUT_DIRECTORY = PROJECT_ROOT / "regional_sales_reports"
DEFAULT_EXCEL_OUTPUT = REPORT_OUTPUT_DIRECTORY / "regional_sales_report.xlsx"
DEFAULT_HTML_OUTPUT = SITE_OUTPUT_DIRECTORY / "index.html"
DEFAULT_COMPANY_DATA = (
    PROJECT_ROOT / "data" / "Company Information.xlsx"
)
DEFAULT_SALES_DATA = (
    PROJECT_ROOT / "data" / "Regional Sales.xlsx"
)

COMPANY_COLUMNS = [
    "company_id",
    "company_name",
    "GICS_sector",
    "GICS_sub_industry",
    "headquarters_loc",
]

REQUIRED_COMPANY_COLUMNS = COMPANY_COLUMNS[:-1]

SALES_COLUMNS = [
    "company_id",
    "gareac",
    "salecs",
    "reporting_date",
]


def parse_arguments():
    parser = argparse.ArgumentParser(
        description="Generate the regional sales report."
    )
    parser.add_argument(
        "start_date",
        help="Start date in YYYY-MM-DD format",
    )
    parser.add_argument(
        "end_date",
        help="End date in YYYY-MM-DD format",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_EXCEL_OUTPUT,
        help="Location of the generated Excel report",
    )
    parser.add_argument(
        "--html-output",
        type=Path,
        default=DEFAULT_HTML_OUTPUT,
        help="Location of the generated browser report",
    )
    parser.add_argument(
        "--company-data",
        type=Path,
        default=DEFAULT_COMPANY_DATA,
        help="Company source data in CSV or Excel format",
    )
    parser.add_argument(
        "--sales-data",
        type=Path,
        default=DEFAULT_SALES_DATA,
        help="Regional sales source data in CSV or Excel format",
    )
    return parser.parse_args()


def validate_dates(start_date_text, end_date_text):
    try:
        start_date = date.fromisoformat(start_date_text)
        end_date = date.fromisoformat(end_date_text)
    except ValueError as error:
        raise ValueError(
            "dates must use YYYY-MM-DD format"
        ) from error

    if start_date > end_date:
        raise ValueError(
            "start date must not be after end date"
        )

    return start_date, end_date


def _read_source_file(path):
    path = Path(path)

    if not path.exists():
        raise ValueError(f"source data file does not exist: {path}")

    if path.suffix.lower() == ".csv":
        return pd.read_csv(path)
    if path.suffix.lower() == ".xlsx":
        return pd.read_excel(path)

    raise ValueError(
        f"source data file must use .csv or .xlsx format: {path}"
    )


def _clean_text(value):
    if pd.isna(value):
        return None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).strip()


def _require_columns(dataframe, columns, source_name):
    missing_columns = [
        column for column in columns if column not in dataframe.columns
    ]

    if missing_columns:
        missing = ", ".join(missing_columns)
        raise ValueError(f"{source_name} is missing required columns: {missing}")


def _require_values(dataframe, columns, source_name):
    for column in columns:
        missing_values = dataframe[column].isna() | (
            dataframe[column].astype(str).str.strip() == ""
        )

        if missing_values.any():
            raise ValueError(
                f"{source_name} is missing values for required column {column}"
            )


def _normalize_company_data(path):
    company_data = _read_source_file(path)

    if company_data.empty:
        raise ValueError("company data contains no records")

    _require_columns(
        company_data,
        REQUIRED_COMPANY_COLUMNS,
        "company data",
    )

    if "headquarters_loc" not in company_data.columns:
        company_data["headquarters_loc"] = None

    company_data = company_data[COMPANY_COLUMNS].copy()

    for column in COMPANY_COLUMNS:
        company_data[column] = company_data[column].map(_clean_text)

    _require_values(
        company_data,
        REQUIRED_COMPANY_COLUMNS,
        "company data",
    )

    return company_data


def _normalize_sales_data(path):
    sales_data = _read_source_file(path)

    if sales_data.empty:
        raise ValueError("sales data contains no records")

    _require_columns(sales_data, SALES_COLUMNS, "sales data")

    sales_data = sales_data[SALES_COLUMNS].copy()

    for column in ("company_id", "gareac"):
        sales_data[column] = sales_data[column].map(_clean_text)

    sales_data["salecs"] = pd.to_numeric(
        sales_data["salecs"],
        errors="coerce",
    )

    reporting_dates = pd.to_datetime(
        sales_data["reporting_date"],
        errors="coerce",
    )

    if reporting_dates.isna().any():
        raise ValueError(
            "sales data contains invalid values for reporting_date"
        )

    sales_data["reporting_date"] = reporting_dates.dt.date.astype(str)

    _require_values(
        sales_data,
        ("company_id", "gareac", "reporting_date"),
        "sales data",
    )

    return sales_data


def _create_database_tables(conn):
    conn.executescript(
        """
        DROP TABLE IF EXISTS regional_sales;
        DROP TABLE IF EXISTS company_info;

        CREATE TABLE company_info (
            company_id TEXT PRIMARY KEY NOT NULL,
            company_name TEXT NOT NULL,
            GICS_sector TEXT NOT NULL,
            GICS_sub_industry TEXT NOT NULL,
            headquarters_loc TEXT
        );

        CREATE TABLE regional_sales (
            company_id TEXT NOT NULL,
            gareac TEXT NOT NULL,
            salecs REAL,
            reporting_date TEXT NOT NULL
        );
        """
    )


def _load_database_tables(conn, company_data_path, sales_data_path):
    company_data = _normalize_company_data(company_data_path)
    sales_data = _normalize_sales_data(sales_data_path)

    _create_database_tables(conn)
    company_data.to_sql(
        "company_info",
        conn,
        if_exists="append",
        index=False,
    )
    sales_data.to_sql(
        "regional_sales",
        conn,
        if_exists="append",
        index=False,
    )
    conn.commit()


def prepare_database(company_data_path, sales_data_path):
    DATABASE_PATH.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(DATABASE_PATH)

    try:
        _load_database_tables(
            conn,
            company_data_path,
            sales_data_path,
        )
    except Exception:
        conn.close()
        raise

    return conn


def retrieve_report_data(conn, start_date, end_date):
    df = pd.read_sql_query(
        """
        SELECT
            company_info.company_name,
            company_info.GICS_sector,
            company_info.GICS_sub_industry,
            regional_sales.gareac AS geographic_area_code,
            SUM(regional_sales.salecs) AS total_sales
        FROM regional_sales
        INNER JOIN company_info
            ON company_info.company_id = regional_sales.company_id
        WHERE regional_sales.reporting_date BETWEEN ? AND ?
        GROUP BY
            company_info.company_name,
            company_info.GICS_sector,
            company_info.GICS_sub_industry,
            regional_sales.gareac
        ORDER BY
            company_info.GICS_sector,
            company_info.GICS_sub_industry,
            company_info.company_name,
            regional_sales.gareac
        """,
        conn,
        params=(start_date.isoformat(), end_date.isoformat()),
    )

    column_names = list(df.columns)
    rows = [
        tuple(row)
        for row in df.where(pd.notna(df), None).itertuples(
            index=False,
            name=None,
        )
    ]

    return column_names, rows


def create_excel_report(
    output_path,
    start_date,
    end_date,
    column_names,
    rows,
    generated_at=None,
):
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if generated_at is None:
        generated_at = datetime.now(timezone.utc).replace(
            microsecond=0,
            tzinfo=None,
        )

    df = pd.DataFrame(rows, columns=column_names)
    df.to_excel(
        output_path,
        sheet_name="Regional Sales Report",
        index=False,
        startrow=6,
    )

    workbook = load_workbook(output_path)
    report_sheet = workbook["Regional Sales Report"]

    report_sheet.merge_cells("A1:E1")
    report_sheet["A1"] = "Regional Sales Report"
    report_sheet["A1"].font = Font(
        bold=True,
        size=16,
        color="FFFFFF",
    )
    report_sheet["A1"].fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    report_sheet["A3"] = "Start Date"
    report_sheet["B3"] = start_date
    report_sheet["C3"] = "End Date"
    report_sheet["D3"] = end_date
    report_sheet["A4"] = "Rows Produced"
    report_sheet["B4"] = len(rows)
    report_sheet["C4"] = "Generated At (UTC)"
    report_sheet["D4"] = generated_at

    for label_cell in ("A3", "C3", "A4", "C4"):
        report_sheet[label_cell].font = Font(bold=True)

    report_sheet["B3"].number_format = "yyyy-mm-dd"
    report_sheet["D3"].number_format = "yyyy-mm-dd"
    report_sheet["D4"].number_format = "yyyy-mm-dd hh:mm:ss"

    header_row = 7
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    for cell in report_sheet[header_row]:
        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )
        cell.fill = header_fill

    report_sheet.freeze_panes = f"A{header_row + 1}"

    column_widths = {
        "A": 28,
        "B": 24,
        "C": 34,
        "D": 24,
        "E": 18,
    }

    for column, width in column_widths.items():
        report_sheet.column_dimensions[column].width = width

    for cell in report_sheet["E"][header_row:]:
        cell.number_format = "#,##0.000"

    if rows:
        last_column = get_column_letter(len(column_names))
        table_reference = (
            f"A{header_row}:{last_column}{report_sheet.max_row}"
        )

        report_table = Table(
            displayName="RegionalSalesReport",
            ref=table_reference,
        )
        report_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        report_sheet.add_table(report_table)

    workbook.save(output_path)


def main():
    arguments = parse_arguments()

    try:
        start_date, end_date = validate_dates(
            arguments.start_date,
            arguments.end_date,
        )

        conn = prepare_database(
            arguments.company_data,
            arguments.sales_data,
        )

        try:
            column_names, rows = retrieve_report_data(
                conn,
                start_date,
                end_date,
            )
        finally:
            conn.close()

        generated_at = datetime.now(timezone.utc).replace(
            microsecond=0,
            tzinfo=None,
        )

        create_excel_report(
            arguments.output,
            start_date,
            end_date,
            column_names,
            rows,
            generated_at,
        )

        browser_excel_output = (
            arguments.html_output.parent / arguments.output.name
        )
        if arguments.output.resolve() != browser_excel_output.resolve():
            browser_excel_output.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(arguments.output, browser_excel_output)

        create_html_report(
            arguments.html_output,
            start_date,
            end_date,
            column_names,
            rows,
            browser_excel_output.name,
            generated_at,
        )

        print(f"Excel report created: {arguments.output}")
        print(f"Web report created: {arguments.html_output}")
        print(f"Rows produced: {len(rows)}")

    except (ValueError, OSError, sqlite3.Error) as error:
        raise SystemExit(str(error)) from error


if __name__ == "__main__":
    main()
