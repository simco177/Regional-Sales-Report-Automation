import csv
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


COMPANY_COLUMNS = (
    "company_id",
    "company_name",
    "GICS_sector",
    "GICS_sub_industry",
    "headquarters_loc",
)

SALES_COLUMNS = (
    "company_id",
    "gareac",
    "salecs",
    "reporting_date",
)


def _string_value(value):
    if value is None:
        return None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).strip()


def _read_csv_rows(path):
    with path.open(newline="", encoding="utf-8-sig") as source_file:
        yield from csv.DictReader(source_file)


def _read_excel_rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)

    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)

        if headers is None:
            return

        header_names = [_string_value(header) for header in headers]

        for values in rows:
            if all(value is None for value in values):
                continue
            yield dict(zip(header_names, values))
    finally:
        workbook.close()


def read_tabular_rows(path):
    path = Path(path)

    if not path.exists():
        raise ValueError(f"source data file does not exist: {path}")

    if path.suffix.lower() == ".csv":
        return _read_csv_rows(path)
    if path.suffix.lower() == ".xlsx":
        return _read_excel_rows(path)

    raise ValueError(
        f"source data file must use .csv or .xlsx format: {path}"
    )


def _required_value(row, possible_columns, source_name):
    for column in possible_columns:
        value = _string_value(row.get(column))
        if value not in (None, ""):
            return value

    expected = " or ".join(possible_columns)
    raise ValueError(
        f"{source_name} is missing a value for required column {expected}"
    )


def normalize_company_rows(path):
    rows = []

    for row in read_tabular_rows(path):
        company_id = _required_value(
            row,
            ("company_id",),
            "company data",
        )
        company_name = _required_value(
            row,
            ("company_name",),
            "company data",
        )
        sector = _required_value(
            row,
            ("GICS_sector",),
            "company data",
        )
        sub_industry = _required_value(
            row,
            ("GICS_sub_industry",),
            "company data",
        )

        rows.append(
            (
                company_id,
                company_name,
                sector,
                sub_industry,
                _string_value(row.get("headquarters_loc")),
            )
        )

    if not rows:
        raise ValueError("company data contains no records")

    return rows


def normalize_sales_rows(path):
    rows = []

    for row in read_tabular_rows(path):
        company_id = _required_value(
            row,
            ("company_id",),
            "sales data",
        )
        region = _required_value(
            row,
            ("gareac",),
            "sales data",
        )
        sales = _required_value(
            row,
            ("salecs",),
            "sales data",
        )
        reporting_date = _required_value(
            row,
            ("reporting_date",),
            "sales data",
        )

        rows.append((company_id, region, sales, reporting_date))

    if not rows:
        raise ValueError("sales data contains no records")

    return rows


def create_source_tables(connection, company_path, sales_path):
    company_rows = normalize_company_rows(company_path)
    sales_rows = normalize_sales_rows(sales_path)

    connection.execute(
        """
        CREATE OR REPLACE TEMP TABLE company_source (
            company_id VARCHAR,
            company_name VARCHAR,
            GICS_sector VARCHAR,
            GICS_sub_industry VARCHAR,
            headquarters_loc VARCHAR
        )
        """
    )
    connection.executemany(
        "INSERT INTO company_source VALUES (?, ?, ?, ?, ?)",
        company_rows,
    )

    connection.execute(
        """
        CREATE OR REPLACE TEMP TABLE sales_source (
            company_id VARCHAR,
            gareac VARCHAR,
            salecs VARCHAR,
            reporting_date VARCHAR
        )
        """
    )
    connection.executemany(
        "INSERT INTO sales_source VALUES (?, ?, ?, ?)",
        sales_rows,
    )
